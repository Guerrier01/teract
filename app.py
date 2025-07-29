# app.py ────────────────────────────────────────────────────────────────
import sys, importlib.util, re
from datetime import datetime
from io import BytesIO

import pandas as pd
import streamlit as st
from openpyxl import Workbook

from style import apply_style
import settings
from columns import BASE_COLUMNS, IA_COLUMNS
from data_io import read_file
from st_utils import preview_df
from llm_utils import build_user_prompt, call_llm
from docs import GUIDE_MD
from xl_utils import ensure_columns
# import notify   # e-mail désactivé

apply_style()

# Sidebar guide
if "show_doc" not in st.session_state:
    st.session_state["show_doc"] = False
st.sidebar.button("📖 Guide utilisateur", on_click=lambda: st.session_state.__setitem__("show_doc", not st.session_state["show_doc"]))
if st.session_state["show_doc"]:
    st.sidebar.markdown(GUIDE_MD, unsafe_allow_html=True)

# Upload
st.markdown("### Chargez votre fichier Teract")
uploaded = st.file_uploader("Formats : xls, xlsm, xlsx, csv", type=["xls", "xlsm", "xlsx", "csv"])

if uploaded:
    try:
        df, wb, ws_entities, groups_row = read_file(uploaded)

        # Nettoyage en-têtes
        df.columns = [re.sub(r"\s+", " ", c.replace("/", " ")).strip() for c in df.columns]

        missing = [c for c in BASE_COLUMNS if c not in df.columns]
        if missing:
            st.markdown(f"<div class='error-msg'>❌ Colonnes manquantes : {', '.join(missing)}</div>", unsafe_allow_html=True)
            st.stop()

        # Ajout IA cols (DataFrame)
        for col in IA_COLUMNS:
            if col not in df.columns:
                df[col] = ""

        st.markdown("<div class='success-msg'>✅ Fichier conforme.</div>", unsafe_allow_html=True)
        preview_df(df, "Aperçu de la feuille Entities", "input_preview")

    except Exception as e:
        st.error(f"Erreur : {e}")
        st.stop()

    # Génération
    if st.button("🚀 Lancer la génération IA"):
        total, errors, total_tokens = len(df), 0, 0
        bar = st.progress(0)
        txt = st.empty()

        # Prépare mapping openpyxl (si wb n'est pas None)
        if wb:
            col_map = ensure_columns(ws_entities, header_row=2, groups_row=1, ia_cols=IA_COLUMNS)

        for idx, row in df.iterrows():
            try:
                r = call_llm(build_user_prompt(row))
                for k in ["Description Marketing Client 1", "Plus produit 1", "Plus produit 2", "Plus produit 3"]:
                    df.at[idx, k] = r["desc"] if k.startswith("Description") else r[k.split()[2].lower()]
                df.at[idx, ["IA DATA", "IAPLUS"]] = 1
                df.at[idx, "Token"] = r["tokens"]
                df.at[idx, "Date"] = datetime.now().strftime("%d/%m/%Y %H:%M")
                df.at[idx, "Commentaires"] = ""
                total_tokens += r["tokens"]

                # Écriture directe dans Excel
                if wb:
                    excel_row = idx + 3  # décalage (headers ligne 2)
                    ws_entities.cell(excel_row, col_map["Description Marketing Client 1"], r["desc"])
                    ws_entities.cell(excel_row, col_map["Plus produit 1"], r["plus1"])
                    ws_entities.cell(excel_row, col_map["Plus produit 2"], r["plus2"])
                    ws_entities.cell(excel_row, col_map["Plus produit 3"], r["plus3"])
                    ws_entities.cell(excel_row, col_map["IA DATA"], 1)
                    ws_entities.cell(excel_row, col_map["Token"], r["tokens"])
                    ws_entities.cell(excel_row, col_map["Date"], datetime.now().strftime("%d/%m/%Y %H:%M"))
                    ws_entities.cell(excel_row, col_map["IAPLUS"], 1)
                    ws_entities.cell(excel_row, col_map["Commentaires"], "")

            except Exception as e:
                errors += 1
                df.at[idx, ["IA DATA", "IAPLUS"]] = 0
                df.at[idx, "Commentaires"] = str(e)[:250]
                if wb:
                    excel_row = idx + 3
                    ws_entities.cell(excel_row, col_map["IA DATA"], 0)
                    ws_entities.cell(excel_row, col_map["IAPLUS"], 0)
                    ws_entities.cell(excel_row, col_map["Commentaires"], str(e)[:250])

            bar.progress((idx + 1) / total)
            txt.markdown(f"<div style='text-align:center;font-weight:600;color:var(--primary);'>{idx+1}/{total} lignes traitées</div>", unsafe_allow_html=True)

        bar.empty(); txt.empty()
        st.success(f"Génération terminée : {total - errors} OK, {errors} erreurs.")

        # Export
        buf = BytesIO()
        if wb:
            wb.save(buf)
            ext = ".xlsm" if uploaded.name.lower().endswith(".xlsm") else ".xlsx"
            mime = "application/vnd.ms-excel" if ext == ".xlsm" else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        else:
            # Cas CSV / Excel simple
            has_xlwt = importlib.util.find_spec("xlwt") and sys.version_info < (3, 12)
            eng, ext, mime = (("xlwt", ".xls", "application/vnd.ms-excel") if has_xlwt else ("openpyxl", ".xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))
            df.to_excel(buf, index=False, engine=eng)
        buf.seek(0)
        st.download_button("💾 Télécharger le fichier", buf, file_name=f"catalogue_enrichi_{datetime.now():%Y%m%d_%H%M}{ext}", mime=mime)

        # notify.send_report(...)  # e-mail désactivé

else:
    st.info("Déposez un fichier pour commencer.")

st.markdown("---\n<div style='text-align:center;font-size:0.85rem;'>© 2025 Teract – Réservé au service Innovation.</div>", unsafe_allow_html=True)
